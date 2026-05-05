import datetime
import json
import math
import re
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MAX_HEADER_SCAN_ROWS = 20


def _is_empty(value):
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def _is_numeric_like(value):
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float)) and not (isinstance(value, float) and math.isnan(value))


def _normalize_for_json(value):
    if _is_empty(value):
        return None
    if isinstance(value, float) and math.isinf(value):
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def _sanitize_identifier(text):
    normalized = re.sub(r"\s+", "_", str(text).strip())
    normalized = re.sub(r"[^A-Za-z0-9_]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "column"


def _make_unique_field_names(header_paths):
    counts = {}
    field_names = []

    for idx, path in enumerate(header_paths, start=1):
        if path:
            base = "__".join(_sanitize_identifier(part) for part in path)
        else:
            base = f"Column_{idx}"

        base = base or f"Column_{idx}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        if count > 1:
            field_names.append(f"{base}_{count}")
        else:
            field_names.append(base)

    return field_names


def _make_unique_labels(labels, fallback_prefix="Column"):
    counts = {}
    unique_labels = []

    for idx, label in enumerate(labels, start=1):
        base = str(label).strip() if label is not None else ""
        if not base:
            base = f"{fallback_prefix}_{idx}"

        count = counts.get(base, 0) + 1
        counts[base] = count
        if count > 1:
            unique_labels.append(f"{base}_{count}")
        else:
            unique_labels.append(base)

    return unique_labels


def _detect_header_depth(rows):
    if not rows:
        return 0

    scan_limit = min(MAX_HEADER_SCAN_ROWS, len(rows))
    fallback_depth = 1

    for row_idx in range(scan_limit):
        row = rows[row_idx]
        non_empty = [value for value in row if not _is_empty(value)]

        if not non_empty:
            if row_idx > 0:
                continue
            return 1

        numeric_count = sum(1 for value in non_empty if _is_numeric_like(value))
        numeric_ratio = numeric_count / len(non_empty)

        # First row that looks like actual data marks the end of header rows.
        if len(non_empty) >= 2 and numeric_ratio >= 0.4:
            return max(1, row_idx)

        fallback_depth = row_idx + 1

    return max(1, fallback_depth)


def _forward_fill_row(values):
    filled = []
    last_seen = None

    for value in values:
        if _is_empty(value):
            filled.append(last_seen)
        else:
            last_seen = str(value).strip()
            filled.append(last_seen)

    return filled


def _clean_header_part(part):
    text = re.sub(r"\s+", " ", str(part)).strip()
    if not text:
        return None
    if re.fullmatch(r"[-_=*./\\|`]+", text):
        return None
    if re.fullmatch(r"\.{2,}\+?", text):
        return None
    return text


def _is_keyed_header_row(row):
    if not row:
        return False
    first = row[0]
    if first is None:
        return False
    first_text = str(first).strip()
    if not first_text.endswith(":"):
        return False

    non_empty = sum(1 for value in row if not _is_empty(value))
    return non_empty >= 3


def _build_header_paths(header_rows):
    if not header_rows:
        return []

    row_count = len(header_rows)
    col_count = len(header_rows[0])
    filled_rows = [_forward_fill_row(row) for row in header_rows]
    keyed_rows = [_is_keyed_header_row(row) for row in filled_rows]

    header_paths = []
    for col_idx in range(col_count):
        parts = []
        for row_idx in range(row_count):
            part = filled_rows[row_idx][col_idx]
            if part is None:
                continue

            if keyed_rows[row_idx]:
                key_name = str(filled_rows[row_idx][0]).strip().rstrip(":").strip().lower()
                if col_idx == 0:
                    continue
                value_text = _clean_header_part(part)
                if not value_text:
                    continue
                part_text = f"{key_name}={value_text}"
            else:
                part_text = _clean_header_part(part)
                if not part_text:
                    continue

            if not parts or parts[-1] != part_text:
                parts.append(part_text)

        header_paths.append(parts)

    return header_paths


def _decompose_header_path(header_path, fallback_label):
    main_key = None
    metadata = {"focus": None, "source": None, "role": None}
    non_key_parts = []
    keyed_qualifiers = []

    for part in header_path:
        if "=" in part:
            key, value = part.split("=", 1)
            cleaned_key = key.strip().lower()
            cleaned_value = value.strip() or None
            if cleaned_key in metadata:
                metadata[cleaned_key] = cleaned_value
            else:
                keyed_qualifiers.append(part)
            continue

        non_key_parts.append(part)

    if non_key_parts:
        main_key = non_key_parts[0]
    else:
        main_key = fallback_label

    return main_key, metadata, non_key_parts, keyed_qualifiers


def _detect_grouped_headers(ws, header_depth, column_count):
    grouped_headers = set()
    spans = []

    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row > header_depth:
            continue
        if merged_range.max_col <= merged_range.min_col:
            continue

        header_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        header_text = _clean_header_part(header_value)
        if not header_text:
            continue

        # Convert from 1-based Excel columns to 0-based JSON column indexes.
        start_idx = max(0, merged_range.min_col - 1)
        end_idx = min(column_count - 1, merged_range.max_col - 1)
        if end_idx <= start_idx:
            continue

        grouped_headers.add(header_text)
        spans.append(
            {
                "header": header_text,
                "excel_range": str(merged_range),
                "start_column_index": start_idx,
                "end_column_index": end_idx,
                "start_column": get_column_letter(merged_range.min_col),
                "end_column": get_column_letter(merged_range.max_col),
                "column_count": (end_idx - start_idx + 1),
            }
        )

    spans.sort(key=lambda item: item["start_column_index"])
    return grouped_headers, spans


def _infer_date_like_columns(field_names, header_paths, data_rows):
    for col_idx, current_name in enumerate(field_names):
        non_empty_values = []
        date_like_count = 0

        for row in data_rows:
            value = row[col_idx] if col_idx < len(row) else None
            if _is_empty(value):
                continue
            non_empty_values.append(value)
            if isinstance(value, (datetime.date, datetime.datetime)):
                date_like_count += 1

        if not non_empty_values:
            continue

        ratio = date_like_count / len(non_empty_values)
        if ratio < 0.7:
            continue

        path = header_paths[col_idx]
        path_joined = " ".join(path).lower()
        if "date" in path_joined:
            continue

        inferred_name = "Date"
        suffix = 2
        while inferred_name in field_names:
            inferred_name = f"Date_{suffix}"
            suffix += 1

        field_names[col_idx] = inferred_name
        header_paths[col_idx] = ["Date"]

    return field_names, header_paths


def _is_visual_spacer_row(values):
    non_empty = [value for value in values if not _is_empty(value)]
    if not non_empty:
        return True

    if len(non_empty) == 1 and isinstance(non_empty[0], str):
        token = non_empty[0].strip()
        if token and re.fullmatch(r"[-_=*./\\|]+", token):
            return True

    return False


def _extract_used_grid(ws):
    max_row = ws.max_row
    max_col = ws.max_column

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
        rows.append(list(row))

    # Trim trailing empty rows
    while rows and all(_is_empty(value) for value in rows[-1]):
        rows.pop()

    if not rows:
        return []

    # Trim trailing empty columns
    rightmost = len(rows[0]) - 1
    while rightmost >= 0 and all(_is_empty(row[rightmost]) for row in rows):
        rightmost -= 1

    if rightmost < 0:
        return []

    return [row[: rightmost + 1] for row in rows]


def _find_active_column_indexes(data_rows, column_count):
    active_indexes = []

    for col_idx in range(column_count):
        has_value = False
        for row in data_rows:
            if _is_visual_spacer_row(row):
                continue
            value = row[col_idx] if col_idx < len(row) else None
            if not _is_empty(value):
                has_value = True
                break

        if has_value:
            active_indexes.append(col_idx)

    if active_indexes:
        return active_indexes

    return list(range(column_count))


def clean_and_convert(excel_file, output_file="output.json"):
    """
    Convert workbook sheets into usable JSON objects with explicit columns and row records.
    """
    print(f"Starting conversion for: {excel_file}...")

    try:
        wb = load_workbook(excel_file, data_only=True)
        output_data = {}
        summary_rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            grid = _extract_used_grid(ws)

            if not grid:
                print(f"Skipping empty sheet: {sheet_name}")
                continue

            raw_row_count = len(grid)
            raw_col_count = len(grid[0])
            raw_non_empty_cells = sum(
                1 for row in grid for value in row if not _is_empty(value)
            )

            header_depth = _detect_header_depth(grid)
            header_rows = grid[:header_depth]
            data_rows = grid[header_depth:]

            header_paths = _build_header_paths(header_rows)
            field_names = _make_unique_field_names(header_paths)
            field_names, header_paths = _infer_date_like_columns(
                field_names, header_paths, data_rows
            )
            grouped_headers, grouped_header_spans = _detect_grouped_headers(
                ws, header_depth, len(header_paths)
            )
            active_column_indexes = _find_active_column_indexes(
                data_rows, len(header_paths)
            )

            columns = []
            metric_label_candidates = []
            for idx in active_column_indexes:
                field_name = field_names[idx]
                path = header_paths[idx]
                display_name = " > ".join(path) if path else f"Column {idx + 1}"
                main_key, row_meta, non_key_parts, keyed_qualifiers = _decompose_header_path(path, display_name)

                # Assign category only when the column falls within the merged-cell span.
                category = None
                for span in grouped_header_spans:
                    if span["start_column_index"] <= idx <= span["end_column_index"]:
                        category = span["header"]
                        break

                # Strip any grouped-header prefix that forward-fill injected into
                # non_key_parts[0], regardless of whether this column is in the span.
                # (forward-fill spreads the merged label into all empty cells to the right)
                grouped_prefix = None
                grouped_header_names = {span["header"] for span in grouped_header_spans}
                if non_key_parts and non_key_parts[0] in grouped_header_names:
                    grouped_prefix = non_key_parts[0]

                non_key_qualifiers = []
                if grouped_prefix is not None:
                    metric_label = non_key_parts[1] if len(non_key_parts) >= 2 else display_name
                    non_key_qualifiers = non_key_parts[2:]
                else:
                    metric_label = main_key
                    if len(non_key_parts) >= 2:
                        non_key_qualifiers = non_key_parts[1:]

                qualifiers = non_key_qualifiers + keyed_qualifiers
                metric_label_candidates.append(metric_label)
                columns.append(
                    {
                        "field_name": field_name,
                        "display_name": display_name,
                        "header_path": path,
                        "source_column_index": idx,
                        "metric_label": metric_label,
                        "category": category,
                        "focus": row_meta["focus"],
                        "source": row_meta["source"],
                        "role": row_meta["role"],
                        "qualifiers": qualifiers,
                    }
                )

            unique_metric_labels = _make_unique_labels(metric_label_candidates, fallback_prefix="Metric")
            for col_idx, metric_label in enumerate(unique_metric_labels):
                columns[col_idx]["metric_label"] = metric_label

            sheet_metric_metadata = {}
            for column in columns:
                metric_label = column["metric_label"]
                entry = {
                    "focus": column["focus"],
                    "source": column["source"],
                    "role": column["role"],
                }
                if column["category"] is not None:
                    entry["category"] = column["category"]
                if column["qualifiers"]:
                    entry["qualifiers"] = column["qualifiers"]
                sheet_metric_metadata[metric_label] = entry

            date_column_index = None
            for col_idx, column in enumerate(columns):
                if column["metric_label"].lower().startswith("date"):
                    date_column_index = col_idx
                    break

            rows = []
            dropped_spacer_rows = 0
            source_data_non_empty = 0
            converted_non_empty = 0

            for source_row_idx, source_row in enumerate(data_rows, start=header_depth + 1):
                if _is_visual_spacer_row(source_row):
                    dropped_spacer_rows += 1
                    continue

                source_data_non_empty += sum(1 for value in source_row if not _is_empty(value))

                date_value = None
                metrics = {}
                row_non_empty = 0
                for col_idx, column in enumerate(columns):
                    source_col_idx = column["source_column_index"]
                    value = source_row[source_col_idx] if source_col_idx < len(source_row) else None
                    normalized = _normalize_for_json(value)
                    if normalized is not None:
                        row_non_empty += 1

                    if date_column_index is not None and col_idx == date_column_index:
                        date_value = normalized
                        continue

                    metric_label = column["metric_label"]
                    metrics[metric_label] = normalized

                if row_non_empty == 0:
                    dropped_spacer_rows += 1
                    continue

                converted_non_empty += row_non_empty
                row_payload = {
                    "_source_row": source_row_idx,
                    "date": date_value,
                    "metrics": metrics,
                }
                rows.append(row_payload)

            integrity_ratio = 1.0
            if source_data_non_empty > 0:
                integrity_ratio = converted_non_empty / source_data_non_empty

            output_data[sheet_name] = {
                "metadata": {
                    "sheet_name": sheet_name,
                    "header_depth": header_depth,
                    "raw_row_count": raw_row_count,
                    "raw_column_count": raw_col_count,
                    "raw_non_empty_cells": raw_non_empty_cells,
                    "source_data_non_empty_cells": source_data_non_empty,
                    "converted_non_empty_cells": converted_non_empty,
                    "dropped_spacer_rows": dropped_spacer_rows,
                    "integrity_ratio": round(integrity_ratio, 4),
                    "grouped_headers": grouped_header_spans,
                },
                "rows": rows,
                "columns": columns,
                "metric_metadata": sheet_metric_metadata,
            }

            summary_rows.append(
                (
                    sheet_name,
                    header_depth,
                    len(columns),
                    len(rows),
                    round(integrity_ratio, 4),
                )
            )

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)

        print(f"Success! Output written to {output_file}")
        print("\nConversion summary:")
        for sheet_name, header_depth, col_count, row_count, ratio in summary_rows:
            warning = " [WARN]" if ratio < 0.9 else ""
            print(
                f"- {sheet_name}: headers={header_depth}, columns={col_count}, "
                f"rows={row_count}, integrity_ratio={ratio}{warning}"
            )

    except FileNotFoundError:
        print(f"Error: The file '{excel_file}' was not found. Please ensure it is in the same directory.")
        sys.exit(1)
    except Exception as exc:
        print(f"An unexpected error occurred: {exc}")
        sys.exit(1)

if __name__ == "__main__":
    # Verbatim reference to the requested file
    input_filename = "Scoreboard Test.xlsx"
    clean_and_convert(input_filename)